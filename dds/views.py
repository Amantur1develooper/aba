from datetime import datetime, time
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from .forms import DDSOpCreateForm, DDSOperationForm, DDSArticleForm
from .models import DDSOperation, DDSArticle, Point, DDSArticle, DDSOperation, CashMovement, CashRegister
from django.db import IntegrityError
from django.shortcuts import redirect
from .forms import DDSOperationForm
from .cash_services import apply_cash_movement
from django.db.models import Sum, Q, F
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from openpyxl import Workbook
from decimal import Decimal, ROUND_HALF_UP
from django.contrib import messages
from django.shortcuts import redirect
from .forms import CashIncassoForm
from .models import CashIncasso
from django.db import transaction
from django.core.exceptions import ValidationError
from .cash_services import apply_cash_movement, FIELD_MAP
from collections import defaultdict
from django.db.models import Sum, Q, F
from django.db.models.functions import Coalesce, TruncDate
from django.shortcuts import redirect, render, get_object_or_404
from django.db.models.functions import Coalesce, TruncDate
from django.db.models import Sum, Q
from .utils import user_hotels_qs
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from .forms import DDSOperationForm
from .models import DDSArticle,DDSArticle, Point, DDSOperation, DDSArticle, CashRegister, CashRegister, DDSOperation, DDSOperation, CashMovement, CashRegister, CashMovement, DDSArticle
from .cash_services import apply_cash_movement, FIELD_MAP
from django.db import transaction
from django.core.exceptions import ValidationError
from .cash_services import apply_cash_movement, FIELD_MAP
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import CashRegister, DDSArticle
from .forms import PointForm
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce, TruncDate
from collections import OrderedDict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ─── Excel helpers ────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # dark blue
SUBHEAD_FILL  = PatternFill("solid", start_color="2E75B6")   # mid blue
INCOME_FILL   = PatternFill("solid", start_color="E2EFDA")   # light green
EXPENSE_FILL  = PatternFill("solid", start_color="FCE4D6")   # light red
TOTAL_FILL    = PatternFill("solid", start_color="D6E4F0")   # pale blue
ZERO_FILL     = PatternFill("solid", start_color="F2F2F2")   # grey

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Calibri", size=10)
TOTAL_FONT   = Font(name="Calibri", bold=True, size=10)
TITLE_FONT   = Font(name="Calibri", bold=True, size=13, color="1F4E79")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

MONEY_FMT = '#,##0.00'
DATE_FMT  = 'DD.MM.YYYY'

_thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _xl_header_row(ws, headers: list, row: int = 1, col_start: int = 1):
    """Write a styled header row."""
    for ci, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER


def _xl_body_row(ws, values: list, row: int, money_cols: set = None, fill=None):
    """Write a styled data row."""
    money_cols = money_cols or set()
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.font      = BODY_FONT
        cell.alignment = RIGHT if ci in money_cols else LEFT
        cell.border    = BORDER
        if isinstance(v, float) and ci in money_cols:
            cell.number_format = MONEY_FMT
        if fill:
            cell.fill = fill


def _xl_total_row(ws, values: list, row: int, money_cols: set = None):
    """Write a styled totals row."""
    money_cols = money_cols or set()
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.font      = TOTAL_FONT
        cell.fill      = TOTAL_FILL
        cell.alignment = RIGHT if ci in money_cols else LEFT
        cell.border    = BORDER
        if isinstance(v, float) and ci in money_cols:
            cell.number_format = MONEY_FMT


def _xl_set_widths(ws, widths: list):
    """Set column widths by list index."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _xl_freeze(ws, row: int = 2, col: int = 1):
    """Freeze panes at given row."""
    ws.freeze_panes = ws.cell(row=row, column=col)


def _xl_autofilter(ws, max_col: int, data_rows: int):
    """Add autofilter to header row."""
    if data_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"


def _xl_info_block(ws, rows: list, start_row: int = 1):
    """Write label/value info block at top of sheet."""
    for i, (label, value) in enumerate(rows, start_row):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font = TOTAL_FONT
        vc.font = BODY_FONT
        lc.fill = vc.fill = PatternFill("solid", start_color="D6E4F0")
        lc.border = vc.border = BORDER
        lc.alignment = LEFT
        vc.alignment = LEFT

# ──────────────────────────────────────────────────────────────────────────────

def _parse_date(d: str):
    try:
        return datetime.strptime(d, "%Y-%m-%d").date()
    except Exception:
        return None



def _day_range(date_obj):
    start = timezone.make_aware(datetime.combine(date_obj, time.min))
    end = timezone.make_aware(datetime.combine(date_obj, time.max))
    return start, end

@login_required
def hotel_detail_export_excel(request, pk):
    hotels_qs = user_hotels_qs(request.user)
    point = get_object_or_404(hotels_qs, pk=pk)

    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to   = _parse_date(request.GET.get("date_to", ""))

    ops = (DDSOperation.objects
           .select_related("article")
           .filter(hotel=point, is_voided=False))

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    rooms_q = (
        Q(source__iexact="rooms") | Q(source__icontains="room") |
        Q(article__name__icontains="номер") |
        Q(article__name__icontains="прожив") |
        Q(article__name__icontains="комнат")
    )
    rooms_ops = ops.filter(article__kind=DDSArticle.INCOME).filter(rooms_q)

    income_total  = ops.filter(article__kind=DDSArticle.INCOME).aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]
    expense_total = ops.filter(article__kind=DDSArticle.EXPENSE).aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]
    balance       = income_total - expense_total

    # Доход по статьям
    by_article = (
        ops.values("article__kind", "article__name")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("article__kind", "-total")
    )

    tz = timezone.get_current_timezone()
    rooms_by_day = (
        rooms_ops
        .annotate(day=TruncDate("happened_at", tzinfo=tz))
        .values("day")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("day")
    )

    wb = Workbook()

    # ── Лист 1: Итоги ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Итоги"
    ws1.sheet_view.showGridLines = False

    # Заголовок отчёта
    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    title_cell.value = f"Отчёт ДДС — {point.name}"
    title_cell.font      = TITLE_FONT
    title_cell.alignment = LEFT
    ws1.row_dimensions[1].height = 22

    # Мета-блок
    period_str = f"{date_from.strftime('%d.%m.%Y') if date_from else 'начало'} → {date_to.strftime('%d.%m.%Y') if date_to else 'конец'}"
    meta = [
        ("Точка", point.name),
        ("Период", period_str),
    ]
    for i, (lbl, val) in enumerate(meta, 3):
        lc = ws1.cell(row=i, column=1, value=lbl)
        vc = ws1.cell(row=i, column=2, value=val)
        lc.font = TOTAL_FONT; vc.font = BODY_FONT
        lc.fill = vc.fill = PatternFill("solid", start_color="D6E4F0")
        lc.border = vc.border = BORDER
        lc.alignment = LEFT; vc.alignment = LEFT

    # Финансовые итоги
    ws1.cell(row=6, column=1, value="").font = BODY_FONT  # spacer
    finance = [
        (7, "💰 Итого приход",  float(income_total),  INCOME_FILL),
        (8, "💸 Итого расход",  float(expense_total), EXPENSE_FILL),
        (9, "📈 Остаток",       float(balance),        TOTAL_FILL),
    ]
    for row, label, value, fill in finance:
        lc = ws1.cell(row=row, column=1, value=label)
        vc = ws1.cell(row=row, column=2, value=value)
        lc.font = vc.font = TOTAL_FONT
        lc.fill = vc.fill = fill
        lc.border = vc.border = BORDER
        lc.alignment = LEFT
        vc.alignment = RIGHT
        vc.number_format = MONEY_FMT

    _xl_set_widths(ws1, [28, 22])

    # ── Лист 2: По статьям ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("📋 По статьям")
    ws2.sheet_view.showGridLines = False
    _xl_header_row(ws2, ["Тип", "Статья", "Сумма (сом)"], row=1)
    row2 = 2
    for r in by_article:
        kind_label = "Доход" if r["article__kind"] == DDSArticle.INCOME else "Расход"
        row_fill   = INCOME_FILL if r["article__kind"] == DDSArticle.INCOME else EXPENSE_FILL
        _xl_body_row(ws2, [kind_label, r["article__name"], float(r["total"])], row=row2, money_cols={3}, fill=row_fill)
        row2 += 1
    _xl_total_row(ws2, ["", "ИТОГО ПРИХОД", float(income_total)], row=row2, money_cols={3})
    row2 += 1
    _xl_total_row(ws2, ["", "ИТОГО РАСХОД", float(expense_total)], row=row2, money_cols={3})
    row2 += 1
    _xl_total_row(ws2, ["", "ОСТАТОК",      float(balance)],       row=row2, money_cols={3})
    _xl_set_widths(ws2, [10, 38, 16])
    _xl_freeze(ws2)
    _xl_autofilter(ws2, 3, row2)

    # ── Лист 3: Номера по дням ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("🛏 Номера по дням")
    ws3.sheet_view.showGridLines = False
    _xl_header_row(ws3, ["Дата", "Доход с номеров (сом)"], row=1)
    row3 = 2
    rooms_total = Decimal("0.00")
    for r in rooms_by_day:
        day_str = r["day"].strftime("%d.%m.%Y") if r["day"] else ""
        _xl_body_row(ws3, [day_str, float(r["total"])], row=row3, money_cols={2})
        rooms_total += r["total"]
        row3 += 1
    if row3 > 2:
        _xl_total_row(ws3, ["ИТОГО", float(rooms_total)], row=row3, money_cols={2})
    _xl_set_widths(ws3, [14, 22])
    _xl_freeze(ws3)

    # ── Лист 4: Все операции ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("📄 Операции")
    ws4.sheet_view.showGridLines = False
    headers4 = ["Дата", "Тип", "Категория / Статья", "Способ оплаты", "Сумма (сом)", "Контрагент", "Источник", "Комментарий"]
    _xl_header_row(ws4, headers4, row=1)
    MONEY_COL = {5}
    row4 = 2
    for op in ops.select_related("article__category").order_by("happened_at"):
        kind = op.article.kind
        fill = INCOME_FILL if kind == DDSArticle.INCOME else EXPENSE_FILL
        article_label = op.article.name
        if op.article.category:
            article_label = f"{op.article.category.name} → {op.article.name}"
        _xl_body_row(ws4, [
            op.happened_at.strftime("%d.%m.%Y %H:%M"),
            op.article.get_kind_display(),
            article_label,
            op.get_method_display(),
            float(op.amount),
            op.counterparty or "",
            op.source or "",
            (op.comment or "")[:500],
        ], row=row4, money_cols=MONEY_COL, fill=fill)
        row4 += 1
    _xl_set_widths(ws4, [16, 9, 38, 14, 14, 24, 12, 40])
    _xl_freeze(ws4)
    _xl_autofilter(ws4, 8, row4 - 2)

    filename = f"tochka_{point.id}_dds.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def unified_report(request):
    # доступ: только superuser/finance_admin
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        # обычному пользователю можно показывать только его отельный дашборд
        return redirect("dds:dds_dashboard")

    hotels_qs = user_hotels_qs(request.user)  # для finance_admin будет "все"
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    ops = DDSOperation.objects.select_related("hotel", "article").filter(is_voided=False, hotel__in=hotels_qs)

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    # Свод по отелям
    by_hotels = (
        ops.values("hotel_id", "hotel__name")
        .annotate(
            income=Coalesce(Sum("amount", filter=Q(article__kind=DDSArticle.INCOME)), Decimal("0.00")),
            expense=Coalesce(Sum("amount", filter=Q(article__kind=DDSArticle.EXPENSE)), Decimal("0.00")),
        )
        .annotate(balance=F("income") - F("expense"))
        .order_by("hotel__name")
    )

    total_income = sum((x["income"] for x in by_hotels), Decimal("0.00"))
    total_expense = sum((x["expense"] for x in by_hotels), Decimal("0.00"))
    total_balance = total_income - total_expense

    # Свод по статьям (по сети)
    by_articles = (
        ops.values("article__kind", "article__name")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("article__kind", "-total")
    )

    return render(
        request,
        "dds/unified_report.html",
        {
            "by_hotels": by_hotels,
            "by_articles": by_articles,
            "total_income": total_income,
            "total_expense": total_expense,
            "total_balance": total_balance,
            "date_from": date_from,
            "date_to": date_to,
        },
    )



def _day_range(date_obj):
    start = timezone.make_aware(datetime.combine(date_obj, time.min))
    end = timezone.make_aware(datetime.combine(date_obj, time.max))
    return start, end


def _user_hotels_qs(user):
    # TODO: адаптируй под свою систему ролей
    if user.is_superuser or getattr(user, "is_finance_admin", False):
        return Point.objects.filter(is_active=True)
    # пример: профиль пользователя хранит отель
    hotel = getattr(getattr(user, "profile", None), "hotel", None)
    return Point.objects.filter(id=hotel.id, is_active=True) if hotel else Point.objects.none()



@login_required
def dds_dashboard(request):
    hotels_qs = _user_hotels_qs(request.user)

    hotel_id = request.GET.get("hotel") or ""
    selected_hotel = None
    if hotel_id:
        selected_hotel = get_object_or_404(hotels_qs, id=hotel_id)

    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    ops = (
        DDSOperation.objects
        .select_related("hotel", "article", "article__category", "article__category__parent")
        .filter(is_voided=False)
    )

    ops = ops.filter(hotel=selected_hotel) if selected_hotel else ops.filter(hotel__in=hotels_qs)

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    # =========================================================
    # Инкассация: НЕ показываем в графиках (и у тебя она ещё убрана из expense_sum)
    # =========================================================
    incasso_cut = Q(article__kind=DDSArticle.EXPENSE) & (
        Q(source__iexact="incasso") | Q(article__name__iexact="Инкассация")
    )
    ops_for_charts = ops.exclude(incasso_cut)

    # -----------------------------
    # Итоги
    # -----------------------------
    income_sum = ops.filter(article__kind=DDSArticle.INCOME).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    # ВАЖНО: сейчас расходы считаются БЕЗ инкассации (как ты и сделал)
    expense_sum = ops_for_charts.filter(article__kind=DDSArticle.EXPENSE).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    balance = income_sum - expense_sum

    # -----------------------------
    # Методы (счет)
    # -----------------------------
    methods = [DDSOperation.CASH, DDSOperation.MKASSA, DDSOperation.ZADATOK, DDSOperation.OPTIMA]
    method_labels = dict(DDSOperation.METHOD_CHOICES)
    method_headers = [{"code": m, "label": method_labels.get(m, m)} for m in methods]

    # =========================================================
    # 1) Таблицы: категории -> подкатегории + разбивка по методам
    # (таблицы строим по ops — инкассация там будет, если не хочешь, скажи)
    # =========================================================
    rows_qs = (
    ops_for_charts.values(
        "article__kind",
        "method",
        "article__category_id",
        "article__category__name",
        "article__category__parent_id",
        "article__category__parent__name",
    )
    .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
)

    rows = list(rows_qs)  # чтобы 2 раза не гонять один и тот же запрос

    def build_groups(kind: str):
        groups_map = {}
        uncat_by_method = defaultdict(lambda: Decimal("0.00"))

        for r in rows:
            if r["article__kind"] != kind:
                continue

            m = r["method"]
            total = r["total"] or Decimal("0.00")

            cat_id = r["article__category_id"]
            cat_name = r["article__category__name"]
            parent_id = r["article__category__parent_id"]
            parent_name = r["article__category__parent__name"]

            if not cat_id:
                uncat_by_method[m] += total
                continue

            if parent_id:
                group_id = parent_id
                group_name = parent_name or "Без названия"
                sub_id = cat_id
                sub_name = cat_name or "Без названия"
            else:
                group_id = cat_id
                group_name = cat_name or "Без названия"
                sub_id = None
                sub_name = None

            g = groups_map.get(group_id)
            if not g:
                g = {
                    "id": group_id,
                    "name": group_name,
                    "total": Decimal("0.00"),
                    "by_method": defaultdict(lambda: Decimal("0.00")),
                    "subs_map": {},
                }
                groups_map[group_id] = g

            g["total"] += total
            g["by_method"][m] += total

            if sub_id:
                s = g["subs_map"].get(sub_id)
                if not s:
                    s = {
                        "id": sub_id,
                        "name": sub_name,
                        "total": Decimal("0.00"),
                        "by_method": defaultdict(lambda: Decimal("0.00")),
                    }
                    g["subs_map"][sub_id] = s
                s["total"] += total
                s["by_method"][m] += total

        groups_list = []
        for g in groups_map.values():
            subs_list = sorted(g["subs_map"].values(), key=lambda x: (x["name"] or "").lower())
            for s in subs_list:
                s["method_totals"] = [s["by_method"][mm] for mm in methods]

            groups_list.append({
                "name": g["name"],
                "total": g["total"],
                "method_totals": [g["by_method"][mm] for mm in methods],
                "subs": subs_list,
            })

        groups_list.sort(key=lambda x: (x["name"] or "").lower())

        uncat = {
            "total": sum(uncat_by_method.values(), Decimal("0.00")),
            "method_totals": [uncat_by_method[mm] for mm in methods],
        }
        return groups_list, uncat

    income_groups, income_uncat = build_groups(DDSArticle.INCOME)
    expense_groups, expense_uncat = build_groups(DDSArticle.EXPENSE)

    # =========================================================
    # 2) Графики: по дням + stack по методам (по ops_for_charts)
    # =========================================================
    tz = timezone.get_current_timezone()
    day_rows_qs = (
        ops_for_charts
        .annotate(day=TruncDate("happened_at", tzinfo=tz))
        .values("day", "article__kind", "method")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("day")
    )
    day_rows = list(day_rows_qs)

    days = sorted({r["day"] for r in day_rows if r["day"] is not None})
    day_labels = [d.strftime("%Y-%m-%d") for d in days]

    grid = {
        DDSArticle.INCOME: {m: defaultdict(lambda: Decimal("0.00")) for m in methods},
        DDSArticle.EXPENSE: {m: defaultdict(lambda: Decimal("0.00")) for m in methods},
    }
    for r in day_rows:
        d = r["day"]
        if not d:
            continue
        kind = r["article__kind"]
        m = r["method"]
        if kind in grid and m in grid[kind]:
            grid[kind][m][d] += (r["total"] or Decimal("0.00"))

    def build_chart(kind: str):
        datasets = []
        for m in methods:
            datasets.append({
                "label": method_labels.get(m, m),
                "data": [float(grid[kind][m][d]) for d in days],
            })
        return {"labels": day_labels, "datasets": datasets}

    income_chart = build_chart(DDSArticle.INCOME)
    expense_chart = build_chart(DDSArticle.EXPENSE)

    # =========================================================
    # 3) Расходы: % по категориям (bar) + pie TOP N (+ Другое)
    #    тоже по ops_for_charts (инкассации там нет)
    # =========================================================
    expense_cat_rows_qs = (
        ops_for_charts.filter(article__kind=DDSArticle.EXPENSE)
        .values(
            "article__category_id",
            "article__category__name",
            "article__category__parent_id",
            "article__category__parent__name",
        )
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
    )
    expense_cat_rows = list(expense_cat_rows_qs)

    cat_totals = defaultdict(lambda: Decimal("0.00"))
    for r in expense_cat_rows:
        total = r["total"] or Decimal("0.00")

        if not r["article__category_id"]:
            group_name = "Без категории"
        else:
            parent_id = r["article__category__parent_id"]
            parent_name = r["article__category__parent__name"]
            cat_name = r["article__category__name"]
            group_name = (parent_name if parent_id else cat_name) or "Без категории"

        cat_totals[group_name] += total

    cat_sorted = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)
    grand_total = sum(cat_totals.values(), Decimal("0.00"))




    expense_cat_percent = {"labels": [], "percent": [], "amounts": [], "grand_total": 0.0}

    if grand_total > 0:
        expense_cat_percent["grand_total"] = float(grand_total)

        for name, total in cat_sorted:
        # точный процент
            pct = (total / grand_total) * Decimal("100")
            pct = pct.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            expense_cat_percent["labels"].append(name)
            expense_cat_percent["percent"].append(float(pct))
            expense_cat_percent["amounts"].append(float(total))

    

    TOP_N = 8
    pie_labels, pie_data = [], []
    other_sum = Decimal("0.00")

    for i, (name, total) in enumerate(cat_sorted):
        if i < TOP_N:
            pie_labels.append(name)
            pie_data.append(float(total))
        else:
            other_sum += total

    if other_sum > 0:
        pie_labels.append("Другое")
        pie_data.append(float(other_sum))

    expense_cat_share = {
        "labels": pie_labels,
        "data": pie_data,
        "grand_total": float(grand_total),  # важно для tooltip в твоём JS
    }
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:dds_dashboard")

    hotels = user_hotels_qs(request.user).order_by("name")

    hotel_id = request.GET.get("hotel")
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    # фильтр по отелю
    hotels_filter = hotels
    if hotel_id:
        hotels_filter = hotels.filter(id=hotel_id)

    

    # 2) ИНКАССАЦИИ
    incassos = CashIncasso.objects.select_related("hotel").filter(hotel__in=hotels_filter)
    return render(request, "dds/dashboard.html", {
        "hotels": hotels_qs,
        "selected_hotel": selected_hotel,
        "date_from": date_from,
        "date_to": date_to,
        "incassos": incassos.order_by("-happened_at")[:300], 
        "income_sum": income_sum,
        "expense_sum": expense_sum,
        "balance": balance,

        "method_headers": method_headers,

        "income_groups": income_groups,
        "income_uncat": income_uncat,
        "expense_groups": expense_groups,
        "expense_uncat": expense_uncat,

        "income_chart": income_chart,
        "expense_chart": expense_chart,

        "expense_cat_percent": expense_cat_percent,
        "expense_cat_share": expense_cat_share,
    })


@login_required
def dds_list(request):
    hotels_qs = _user_hotels_qs(request.user)

    ops = DDSOperation.objects.select_related("hotel", "article").filter(hotel__in=hotels_qs)

    hotel_id = request.GET.get("hotel")
    kind = request.GET.get("kind")  # income/expense
    article_id = request.GET.get("article")
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    if hotel_id:
        ops = ops.filter(hotel_id=hotel_id)
    if kind in (DDSArticle.INCOME, DDSArticle.EXPENSE):
        ops = ops.filter(article__kind=kind)
    if article_id:
        ops = ops.filter(article_id=article_id)
    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    articles = DDSArticle.objects.filter(is_active=True)

    return render(
        request,
        "dds/operation_list.html",
        {
            "ops": ops[:500],  # на MVP
            "hotels": hotels_qs,
            "articles": articles,
            "filters": {"hotel": hotel_id, "kind": kind, "article": article_id, "date_from": date_from, "date_to": date_to},
        },
    )




def _is_rooms_income(op) -> bool:
    src = (op.source or "").lower()
    name = (op.article.name or "").lower()
    return (
        src == "rooms"
        or "room" in src
        or ("номер" in name)
        or ("прожив" in name)
        or ("комнат" in name)
    )


@login_required
def dds_op_add(request, hotel_id: int, kind: str):
    hotels = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels, id=hotel_id)

    register, _ = CashRegister.objects.get_or_create(hotel=hotel)

    category_id = request.GET.get("category") or request.POST.get("category") or None

    if request.method == "POST":
        form = DDSOpCreateForm(
            request.POST,
            kind=kind,
            category_id=category_id,
            hotel=hotel,   # ✅ ВАЖНО: без этого будет показывать всё
        )

        if form.is_valid():
            op = form.save(commit=False)
            op.hotel = hotel
            op.created_by = request.user

            direction = CashMovement.IN if op.article.kind == DDSArticle.INCOME else CashMovement.OUT
            is_incasso = (op.source or "").lower() == "incasso"

            try:
                with transaction.atomic():
                    register = CashRegister.objects.select_for_update().get(pk=register.pk)

                    if (not is_incasso) and direction == CashMovement.OUT:
                        field = FIELD_MAP.get(op.method)
                        if not field:
                            messages.error(request, "Неверный способ оплаты/счет.")
                            return render(request, "dds/dds_quick_op_form.html", {
                                "hotel": hotel, "register": register, "kind": kind,
                                "form": form, "category_id": category_id or "",
                            })

                        current = getattr(register, field) or Decimal("0.00")
                        if op.amount > current:
                            messages.error(
                                request,
                                f"Недостаточно средств на счете {op.get_method_display()}. Доступно: {current}"
                            )
                            return render(request, "dds/dds_quick_op_form.html", {
                                "hotel": hotel, "register": register, "kind": kind,
                                "form": form, "category_id": category_id or "",
                            })

                    op.save()

                    if not is_incasso:
                        exists = CashMovement.objects.filter(
                            dds_operation=op,
                            account=op.method,
                            direction=direction,
                        ).exists()

                        if not exists:
                            try:
                                apply_cash_movement(
                                    hotel=hotel,
                                    account=op.method,
                                    direction=direction,
                                    amount=op.amount,
                                    created_by=request.user,
                                    happened_at=op.happened_at,
                                    comment=op.comment,
                                    dds_operation=op,
                                )
                            except IntegrityError:
                                pass

            except ValidationError as e:
                messages.error(request, str(e))
                return render(request, "dds/dds_quick_op_form.html", {
                    "hotel": hotel, "register": register, "kind": kind,
                    "form": form, "category_id": category_id or "",
                })

            messages.success(request, "Операция сохранена и касса обновлена.")
            return redirect("dds:hotel_detail", hotel.id)

        messages.error(request, "Исправьте ошибки в форме.")

    else:
        form = DDSOpCreateForm(
            kind=kind,
            category_id=category_id,
            hotel=hotel,   # ✅ ВАЖНО
        )

    return render(request, "dds/dds_quick_op_form.html", {
        "hotel": hotel,
        "register": register,
        "kind": kind,
        "form": form,
        "category_id": category_id or "",
    })

@login_required
def dds_create(request):
    hotels_qs = user_hotels_qs(request.user)
    if not hotels_qs.exists():
        messages.error(request, "У вас не назначен отель. Обратитесь к администратору.")
        return redirect("dds:dds_list")

    only_hotel = hotels_qs.first() if hotels_qs.count() == 1 else None

    selected_hotel = None
    if only_hotel:
        selected_hotel = only_hotel
    else:
        hotel_id = request.POST.get("hotel") if request.method == "POST" else request.GET.get("hotel")
        if hotel_id and hotels_qs.filter(id=hotel_id).exists():
            selected_hotel = hotels_qs.get(id=hotel_id)

    kind = request.GET.get("kind") or request.POST.get("kind")
    if kind not in (DDSArticle.INCOME, DDSArticle.EXPENSE):
        kind = None

    if request.method == "POST":
        form = DDSOperationForm(request.POST, hotel=selected_hotel, kind=kind)
        form.fields["hotel"].queryset = hotels_qs


        if only_hotel:
            form.fields["hotel"].initial = only_hotel
            form.fields["hotel"].disabled = True

        # ❌ НЕ ДЕЛАЙ ТАК:
        # form.fields["article"].queryset = DDSArticle.objects.filter(is_active=True)

        if form.is_valid():
            op = form.save(commit=False)
            
            
            if only_hotel:
            
                op.hotel = only_hotel
            op.created_by = request.user
            op.save()

            messages.success(request, "Операция добавлена.")
            return redirect("dds:hotel_detail", pk=op.hotel_id)

    else:
        form = DDSOperationForm(hotel=selected_hotel, kind=kind)



        form.fields["hotel"].queryset = hotels_qs




        if selected_hotel:
            form.fields["hotel"].initial = selected_hotel
        if only_hotel:
            form.fields["hotel"].initial = only_hotel
            form.fields["hotel"].disabled = True

    
    reg = None
    
    if selected_hotel:
        reg, _ = CashRegister.objects.get_or_create(hotel=selected_hotel)

    return render(request, "dds/operation_form.html", {
        "form": form,
        "reg": reg,
        "selected_hotel": selected_hotel,
        "kind": kind,
    })





@login_required
def dds_void(request, pk):
    hotels_qs = _user_hotels_qs(request.user)
    op = get_object_or_404(DDSOperation, pk=pk, hotel__in=hotels_qs)

    if request.method == "POST":
        reason = (request.POST.get("reason") or "").strip()
        if not reason:
            messages.error(request, "Укажи причину сторно.")
            return redirect("dds:dds_list")
        op.void(request.user, reason)
        messages.success(request, "Операция отменена (сторно).")
        return redirect("dds:dds_list")

    return render(request, "dds/void_confirm.html", {"op": op})


@login_required
def dds_articles(request):
    # TODO: доступ только бухгалтеру/админу
    if not (request.user.is_superuser or getattr(request.user, "is_finance_admin", False)):
        return redirect("dds:dds_dashboard")

    if request.method == "POST":
        form = DDSArticleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Статья сохранена.")
            return redirect("dds:dds_articles")
    else:
        form = DDSArticleForm()

    articles = DDSArticle.objects.all()
    return render(request, "dds/articles.html", {"form": form, "articles": articles})





@login_required
def hotel_catalog(request):
    # каталог менять/добавлять — только админ/финанс
    is_fin_admin = request.user.is_superuser or getattr(getattr(request.user, "profile", None), "is_finance_admin", False)

    # что показываем:
    hotels = Point.objects.all().order_by("name") if is_fin_admin else user_hotels_qs(request.user).order_by("name")

    form = None
    if is_fin_admin:
        form = PointForm(request.POST or None)
        if request.method == "POST" and form.is_valid():
            form.save()
            messages.success(request, "Отель добавлен.")
            return redirect("dds:hotel_catalog")

    return render(request, "dds/hotel_catalog.html", {"hotels": hotels, "form": form, "is_fin_admin": is_fin_admin})




@login_required
def hotel_list(request):
    hotels = list(user_hotels_qs(request.user).order_by("name"))

    # Берём кассы одним запросом и мапим по hotel_id
    registers_by_hotel = CashRegister.objects.filter(hotel__in=hotels).in_bulk(field_name="hotel_id")

    # приклеиваем register к каждому отелю (может быть None)
    for h in hotels:
        h.register = registers_by_hotel.get(h.id)

    return render(request, "dds/hotel_list.html", {"hotels": hotels})


@login_required
def hotel_detail(request, pk):
    hotels_qs = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels_qs, pk=pk)

    reg, _ = CashRegister.objects.get_or_create(hotel=hotel)

    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    ops = (
        DDSOperation.objects
        .select_related(
            "article",
            "article__category",
            "article__category__parent",
        )
        .filter(hotel=hotel, is_voided=False)
    )

    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    # ✅ Итоги ДДС за период
    income_total = ops.filter(article__kind=DDSArticle.INCOME).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    expense_total = ops.filter(article__kind=DDSArticle.EXPENSE).aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    balance = income_total - expense_total

    # ✅ Доход с номеров
    rooms_q = (
        Q(source__iexact="rooms") |
        Q(source__icontains="room") |
        Q(article__name__icontains="номер") |
        Q(article__name__icontains="прожив") |
        Q(article__name__icontains="комнат")
    )
    rooms_ops = ops.filter(article__kind=DDSArticle.INCOME).filter(rooms_q)

    rooms_income_total = rooms_ops.aggregate(
        s=Coalesce(Sum("amount"), Decimal("0.00"))
    )["s"]

    tz = timezone.get_current_timezone()
    rooms_by_day = (
        rooms_ops
        .annotate(day=TruncDate("happened_at", tzinfo=tz))
        .values("day")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("day")
    )

    last_ops = ops.order_by("-happened_at")[:50]

    # ✅ ДДС по счетам за период (как у тебя)
    methods = [DDSOperation.CASH, DDSOperation.MKASSA, DDSOperation.ZADATOK, DDSOperation.OPTIMA]
    method_labels = dict(DDSOperation.METHOD_CHOICES)

    period_rows = []
    for m in methods:
        inc = ops.filter(article__kind=DDSArticle.INCOME, method=m).aggregate(
            s=Coalesce(Sum("amount"), Decimal("0.00"))
        )["s"]
        exp = ops.filter(article__kind=DDSArticle.EXPENSE, method=m).aggregate(
            s=Coalesce(Sum("amount"), Decimal("0.00"))
        )["s"]
        period_rows.append({
            "code": m,
            "label": method_labels.get(m, m),
            "income": inc,
            "expense": exp,
            "delta": inc - exp,
        })

    # ==========================================================
    # ✅ НОВОЕ: Доходы/Расходы по категориям и подкатегориям
    # ==========================================================
    def build_cat_groups(kind: str):
        """
        Возвращает:
          groups: [
            {id, name, total, subs:[{id,name,total}, ...]},
            ...
          ]
          uncategorized_total: Decimal
        """
        rows = (
            ops.filter(article__kind=kind)
            .values(
                "article__category_id",
                "article__category__name",
                "article__category__parent_id",
                "article__category__parent__name",
            )
            .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        )

        # group_id = parent_id если есть, иначе category_id (верхний уровень)
        groups = {}  # group_id -> dict
        uncategorized_total = Decimal("0.00")

        for r in rows:
            cat_id = r["article__category_id"]
            cat_name = r["article__category__name"]
            parent_id = r["article__category__parent_id"]
            parent_name = r["article__category__parent__name"]
            total = r["total"] or Decimal("0.00")

            if not cat_id:
                uncategorized_total += total
                continue

            if parent_id:
                # подкатегория -> складываем в родителя
                group_id = parent_id
                group_name = parent_name or "Без названия"
                g = groups.get(group_id)
                if not g:
                    g = {"id": group_id, "name": group_name, "total": Decimal("0.00"), "subs": []}
                    groups[group_id] = g

                g["subs"].append({"id": cat_id, "name": cat_name or "Без названия", "total": total})
                g["total"] += total
            else:
                # верхняя категория
                group_id = cat_id
                group_name = cat_name or "Без названия"
                g = groups.get(group_id)
                if not g:
                    g = {"id": group_id, "name": group_name, "total": Decimal("0.00"), "subs": []}
                    groups[group_id] = g
                # если операции записаны прямо на верхнюю категорию — добавляем в total
                g["total"] += total

        # сортировка
        groups_list = sorted(groups.values(), key=lambda x: (x["name"] or "").lower())
        for g in groups_list:
            g["subs"] = sorted(g["subs"], key=lambda x: (x["name"] or "").lower())

        return groups_list, uncategorized_total

    income_groups, income_uncat = build_cat_groups(DDSArticle.INCOME)
    expense_groups, expense_uncat = build_cat_groups(DDSArticle.EXPENSE)

    from .models import PointContact
    contacts = PointContact.objects.filter(point=hotel)

    return render(request, "dds/hotel_detail.html", {
        "hotel": hotel,
        "reg": reg,
        "contacts": contacts,

        # фильтр периода
        "date_from": date_from,
        "date_to": date_to,

        # итоги
        "income_total": income_total,
        "expense_total": expense_total,
        "balance": balance,

        # номера
        "rooms_income_total": rooms_income_total,
        "rooms_by_day": rooms_by_day,

        # по счетам за период
        "period_rows": period_rows,

        # последние операции
        "last_ops": last_ops,

        # ✅ категории/подкатегории
        "income_groups": income_groups,
        "income_uncat": income_uncat,
        "expense_groups": expense_groups,
        "expense_uncat": expense_uncat,
    })




@login_required
def unified_report_export_excel(request):
    # ✅ доступ только superuser/finance_admin
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:dds_dashboard")

    hotels_qs = user_hotels_qs(request.user)

    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to   = _parse_date(request.GET.get("date_to", ""))

    ops = DDSOperation.objects.select_related("hotel", "article", "article__category").filter(
        is_voided=False,
        hotel__in=hotels_qs,
    )
    if date_from:
        start, _ = _day_range(date_from)
        ops = ops.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        ops = ops.filter(happened_at__lte=end)

    # Свод по точкам
    by_points = (
        ops.values("hotel_id", "hotel__name")
        .annotate(
            income=Coalesce(Sum("amount", filter=Q(article__kind=DDSArticle.INCOME)), Decimal("0.00")),
            expense=Coalesce(Sum("amount", filter=Q(article__kind=DDSArticle.EXPENSE)), Decimal("0.00")),
        )
        .annotate(balance=F("income") - F("expense"))
        .order_by("hotel__name")
    )
    by_points_list = list(by_points)  # evaluate once

    total_income  = sum((x["income"]  for x in by_points_list), Decimal("0.00"))
    total_expense = sum((x["expense"] for x in by_points_list), Decimal("0.00"))
    total_balance = total_income - total_expense

    # Свод по статьям
    by_articles = (
        ops.values("article__kind", "article__category__name", "article__name")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("article__kind", "article__category__name", "-total")
    )

    # Детальные операции
    all_ops = ops.order_by("hotel__name", "happened_at")

    wb = Workbook()

    # ── Лист 1: Итоги ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Итоги"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:C1")
    tc = ws1["A1"]
    tc.value = "Единый отчёт ДДС по сети точек"
    tc.font = TITLE_FONT; tc.alignment = LEFT
    ws1.row_dimensions[1].height = 24

    period_str = f"{date_from.strftime('%d.%m.%Y') if date_from else 'начало'} → {date_to.strftime('%d.%m.%Y') if date_to else 'конец'}"
    for i, (lbl, val) in enumerate([("Период", period_str), ("Кол-во точек", len(by_points_list))], 3):
        lc = ws1.cell(row=i, column=1, value=lbl)
        vc = ws1.cell(row=i, column=2, value=val)
        lc.font = TOTAL_FONT; vc.font = BODY_FONT
        lc.fill = vc.fill = PatternFill("solid", start_color="D6E4F0")
        lc.border = vc.border = BORDER
        lc.alignment = LEFT; vc.alignment = LEFT

    for row, label, value, fill in [
        (6,  "💰 Итого приход по сети",  float(total_income),  INCOME_FILL),
        (7,  "💸 Итого расход по сети",  float(total_expense), EXPENSE_FILL),
        (8,  "📈 Остаток по сети",       float(total_balance), TOTAL_FILL),
    ]:
        lc = ws1.cell(row=row, column=1, value=label)
        vc = ws1.cell(row=row, column=2, value=value)
        lc.font = vc.font = TOTAL_FONT
        lc.fill = vc.fill = fill
        lc.border = vc.border = BORDER
        lc.alignment = LEFT; vc.alignment = RIGHT
        vc.number_format = MONEY_FMT
    _xl_set_widths(ws1, [32, 20])

    # ── Лист 2: По точкам ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("📍 По точкам")
    ws2.sheet_view.showGridLines = False
    _xl_header_row(ws2, ["Точка", "Приход (сом)", "Расход (сом)", "Остаток (сом)"], row=1)
    for ri, r in enumerate(by_points_list, 2):
        bal = float(r["balance"])
        bal_fill = INCOME_FILL if bal >= 0 else EXPENSE_FILL
        _xl_body_row(ws2, [r["hotel__name"], float(r["income"]), float(r["expense"]), bal],
                     row=ri, money_cols={2, 3, 4})
        # цвет остатка
        ws2.cell(row=ri, column=4).fill = bal_fill
    total_row = len(by_points_list) + 2
    _xl_total_row(ws2, ["ИТОГО", float(total_income), float(total_expense), float(total_balance)],
                  row=total_row, money_cols={2, 3, 4})
    _xl_set_widths(ws2, [30, 16, 16, 16])
    _xl_freeze(ws2)
    _xl_autofilter(ws2, 4, len(by_points_list))

    # ── Лист 3: По статьям ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("📋 По статьям")
    ws3.sheet_view.showGridLines = False
    _xl_header_row(ws3, ["Тип", "Категория", "Статья", "Сумма (сом)"], row=1)
    ri3 = 2
    inc_total = exp_total = Decimal("0.00")
    for r in by_articles:
        kind_label = "Доход" if r["article__kind"] == DDSArticle.INCOME else "Расход"
        row_fill   = INCOME_FILL if r["article__kind"] == DDSArticle.INCOME else EXPENSE_FILL
        cat = r.get("article__category__name") or "—"
        _xl_body_row(ws3, [kind_label, cat, r["article__name"], float(r["total"])],
                     row=ri3, money_cols={4}, fill=row_fill)
        if r["article__kind"] == DDSArticle.INCOME:
            inc_total += r["total"]
        else:
            exp_total += r["total"]
        ri3 += 1
    _xl_total_row(ws3, ["", "", "ИТОГО ДОХОД",  float(inc_total)], row=ri3, money_cols={4}); ri3 += 1
    _xl_total_row(ws3, ["", "", "ИТОГО РАСХОД", float(exp_total)], row=ri3, money_cols={4}); ri3 += 1
    _xl_total_row(ws3, ["", "", "ОСТАТОК",      float(inc_total - exp_total)], row=ri3, money_cols={4})
    _xl_set_widths(ws3, [10, 24, 36, 16])
    _xl_freeze(ws3)
    _xl_autofilter(ws3, 4, ri3 - 2)

    # ── Лист 4: Все операции ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("📄 Все операции")
    ws4.sheet_view.showGridLines = False
    _xl_header_row(ws4, ["Дата", "Точка", "Тип", "Категория", "Статья", "Способ оплаты",
                          "Сумма (сом)", "Контрагент", "Источник", "Комментарий"], row=1)
    ri4 = 2
    for op in all_ops:
        fill = INCOME_FILL if op.article.kind == DDSArticle.INCOME else EXPENSE_FILL
        cat = op.article.category.name if op.article.category else "—"
        _xl_body_row(ws4, [
            op.happened_at.strftime("%d.%m.%Y %H:%M"),
            op.hotel.name,
            op.article.get_kind_display(),
            cat,
            op.article.name,
            op.get_method_display(),
            float(op.amount),
            op.counterparty or "",
            op.source or "",
            (op.comment or "")[:500],
        ], row=ri4, money_cols={7}, fill=fill)
        ri4 += 1
    _xl_set_widths(ws4, [16, 20, 9, 22, 30, 14, 14, 24, 12, 40])
    _xl_freeze(ws4)
    _xl_autofilter(ws4, 10, ri4 - 2)

    filename = "edinyi_otchet_dds.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def incasso_create(request, pk):
    hotels_qs = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels_qs, pk=pk)

    reg, _ = CashRegister.objects.get_or_create(hotel=hotel)

    if request.method == "POST":
        form = CashIncassoForm(request.POST)
        if form.is_valid():
            inc = form.save(commit=False)
            inc.hotel = hotel
            inc.created_by = request.user

            try:
                with transaction.atomic():
                    reg = CashRegister.objects.select_for_update().get(pk=reg.pk)

                    field = FIELD_MAP.get(inc.method)
                    current = getattr(reg, field) if field else None
                    current = current or Decimal("0.00")

                    if inc.amount > current:
                        messages.error(
                            request,
                            f"Недостаточно средств на счете {inc.get_method_display()}. Доступно: {current}"
                        )
                        return render(request, "dds/incasso_form.html", {"form": form, "hotel": hotel, "reg": reg})

                    inc.save()

                    article, _ = DDSArticle.objects.get_or_create(
                        kind=DDSArticle.EXPENSE,
                        name="Инкассация",
                        defaults={"is_active": True},
                    )

                    op_incasso = DDSOperation.objects.create(
                        hotel=hotel,
                        article=article,
                        amount=inc.amount,
                        happened_at=inc.happened_at,
                        method=inc.method,      # ✅ с какого счета забрали
                        counterparty="Бухгалтерия",
                        comment=inc.comment,
                        source="incasso",
                        created_by=request.user,
                    )

                    apply_cash_movement(
                        hotel=hotel,
                        account=inc.method,         # ✅ списываем именно отсюда
                        direction=CashMovement.OUT,
                        amount=inc.amount,
                        created_by=request.user,
                        happened_at=inc.happened_at,
                        comment=inc.comment,
                        dds_operation=op_incasso,
                        incasso=inc,
                    )

            except ValidationError as e:
                messages.error(request, str(e))
                return render(request, "dds/incasso_form.html", {"form": form, "hotel": hotel, "reg": reg})

            messages.success(request, "Инкассация создана. Средства списаны.")
            return redirect("dds:hotel_detail", pk=hotel.id)
    else:
        form = CashIncassoForm()

    return render(request, "dds/incasso_form.html", {"form": form, "hotel": hotel, "reg": reg})


@login_required
def accounting(request):
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:dds_dashboard")

    hotels = user_hotels_qs(request.user).order_by("name")

    hotel_id = request.GET.get("hotel")
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to = _parse_date(request.GET.get("date_to", ""))

    # фильтр по отелю
    hotels_filter = hotels
    if hotel_id:
        hotels_filter = hotels.filter(id=hotel_id)

    # 1) РАСХОДЫ (не инкассация!)
    expenses = DDSOperation.objects.select_related("hotel", "article").filter(
        is_voided=False,
        hotel__in=hotels_filter,
        article__kind=DDSArticle.EXPENSE,
    ).exclude(source="incasso")

    # 2) ИНКАССАЦИИ
    incassos = CashIncasso.objects.select_related("hotel").filter(hotel__in=hotels_filter)

    # фильтр по датам
    if date_from:
        start, _ = _day_range(date_from)
        expenses = expenses.filter(happened_at__gte=start)
        incassos = incassos.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        expenses = expenses.filter(happened_at__lte=end)
        incassos = incassos.filter(happened_at__lte=end)

    expense_total = expenses.aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]
    incasso_total = incassos.aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]

    return render(request, "dds/accounting.html", {
        "hotels": hotels,
        "selected_hotel": hotel_id,
        "date_from": date_from,
        "date_to": date_to,
        "expenses": expenses.order_by("-happened_at")[:300],
        "incassos": incassos.order_by("-happened_at")[:300],
        "expense_total": expense_total,
        "incasso_total": incasso_total,
    })


@login_required
def accounting_export_excel(request):
    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:dds_dashboard")

    hotels = user_hotels_qs(request.user)

    hotel_id  = request.GET.get("hotel")
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to   = _parse_date(request.GET.get("date_to", ""))

    hotels_filter = hotels
    if hotel_id:
        hotels_filter = hotels.filter(id=hotel_id)

    expenses = DDSOperation.objects.select_related("hotel", "article", "article__category").filter(
        is_voided=False,
        hotel__in=hotels_filter,
        article__kind=DDSArticle.EXPENSE,
    ).exclude(source="incasso")

    incassos = CashIncasso.objects.select_related("hotel").filter(hotel__in=hotels_filter)

    if date_from:
        start, _ = _day_range(date_from)
        expenses = expenses.filter(happened_at__gte=start)
        incassos = incassos.filter(happened_at__gte=start)
    if date_to:
        _, end = _day_range(date_to)
        expenses = expenses.filter(happened_at__lte=end)
        incassos = incassos.filter(happened_at__lte=end)

    exp_total = expenses.aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]
    inc_total = incassos.aggregate(s=Coalesce(Sum("amount"), Decimal("0.00")))["s"]

    # Свод расходов по статьям
    exp_by_article = (
        expenses.values("article__category__name", "article__name")
        .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("article__category__name", "-total")
    )

    wb = Workbook()

    # ── Лист 1: Итоги ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Итоги"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:B1")
    tc = ws1["A1"]
    tc.value = "Бухгалтерский отчёт"
    tc.font = TITLE_FONT; tc.alignment = LEFT
    ws1.row_dimensions[1].height = 22

    period_str = f"{date_from.strftime('%d.%m.%Y') if date_from else 'начало'} → {date_to.strftime('%d.%m.%Y') if date_to else 'конец'}"
    point_label = hotel_id or "Все точки"
    for i, (lbl, val) in enumerate([("Период", period_str), ("Точка", point_label)], 3):
        lc = ws1.cell(row=i, column=1, value=lbl)
        vc = ws1.cell(row=i, column=2, value=val)
        lc.font = TOTAL_FONT; vc.font = BODY_FONT
        lc.fill = vc.fill = PatternFill("solid", start_color="D6E4F0")
        lc.border = vc.border = BORDER
        lc.alignment = LEFT; vc.alignment = LEFT

    for row, label, value, fill in [
        (6, "💸 Расходы (без инкассации)", float(exp_total), EXPENSE_FILL),
        (7, "🏦 Инкассации",               float(inc_total), TOTAL_FILL),
        (8, "📊 Итого выбыло",             float(exp_total + inc_total), PatternFill("solid", start_color="F8CBAD")),
    ]:
        lc = ws1.cell(row=row, column=1, value=label)
        vc = ws1.cell(row=row, column=2, value=value)
        lc.font = vc.font = TOTAL_FONT
        lc.fill = vc.fill = fill
        lc.border = vc.border = BORDER
        lc.alignment = LEFT; vc.alignment = RIGHT
        vc.number_format = MONEY_FMT
    _xl_set_widths(ws1, [30, 22])

    # ── Лист 2: Свод расходов по статьям ──────────────────────────────────────
    ws2 = wb.create_sheet("📋 Свод расходов")
    ws2.sheet_view.showGridLines = False
    _xl_header_row(ws2, ["Категория", "Статья расхода", "Сумма (сом)"], row=1)
    ri2 = 2
    for r in exp_by_article:
        cat = r.get("article__category__name") or "—"
        _xl_body_row(ws2, [cat, r["article__name"], float(r["total"])],
                     row=ri2, money_cols={3}, fill=EXPENSE_FILL)
        ri2 += 1
    _xl_total_row(ws2, ["", "ИТОГО РАСХОДЫ", float(exp_total)], row=ri2, money_cols={3})
    _xl_set_widths(ws2, [24, 36, 16])
    _xl_freeze(ws2)
    _xl_autofilter(ws2, 3, ri2 - 2)

    # ── Лист 3: Расходы детально ───────────────────────────────────────────────
    ws3 = wb.create_sheet("💸 Расходы")
    ws3.sheet_view.showGridLines = False
    _xl_header_row(ws3, ["Дата", "Точка", "Категория", "Статья", "Способ оплаты",
                          "Сумма (сом)", "Контрагент", "Комментарий"], row=1)
    ri3 = 2
    for op in expenses.order_by("happened_at"):
        cat = op.article.category.name if op.article.category else "—"
        _xl_body_row(ws3, [
            op.happened_at.strftime("%d.%m.%Y %H:%M"),
            op.hotel.name,
            cat,
            op.article.name,
            op.get_method_display(),
            float(op.amount),
            op.counterparty or "",
            (op.comment or "")[:500],
        ], row=ri3, money_cols={6}, fill=EXPENSE_FILL)
        ri3 += 1
    if ri3 > 2:
        _xl_total_row(ws3, ["", "", "", "", "ИТОГО", float(exp_total), "", ""], row=ri3, money_cols={6})
    _xl_set_widths(ws3, [16, 20, 22, 30, 14, 14, 24, 40])
    _xl_freeze(ws3)
    _xl_autofilter(ws3, 8, ri3 - 2)

    # ── Лист 4: Инкассации ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("🏦 Инкассации")
    ws4.sheet_view.showGridLines = False
    _xl_header_row(ws4, ["Дата", "Точка", "Счёт списания", "Сумма (сом)", "Комментарий", "Создал"], row=1)
    ri4 = 2
    for inc in incassos.order_by("happened_at"):
        _xl_body_row(ws4, [
            inc.happened_at.strftime("%d.%m.%Y %H:%M"),
            inc.hotel.name,
            inc.get_method_display(),
            float(inc.amount),
            (inc.comment or "")[:500],
            getattr(inc.created_by, "username", ""),
        ], row=ri4, money_cols={4})
        ri4 += 1
    if ri4 > 2:
        _xl_total_row(ws4, ["", "", "ИТОГО", float(inc_total), "", ""], row=ri4, money_cols={4})
    _xl_set_widths(ws4, [16, 20, 16, 14, 40, 16])
    _xl_freeze(ws4)
    _xl_autofilter(ws4, 6, ri4 - 2)

    filename = "bukhgalteriya_otchet.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def global_cash_view(request):
    from .models import GlobalCashRegister, GlobalCashOperation
    from .cash_services import global_cash_income, global_cash_expense

    profile = getattr(request.user, "profile", None)
    if not (request.user.is_superuser or (profile and profile.is_finance_admin)):
        return redirect("dds:hotel_list")

    gcr = GlobalCashRegister.get()
    ops = GlobalCashOperation.objects.prefetch_related("distributions__point").order_by("-happened_at")[:100]

    ACCOUNT_CHOICES = [
        ("cash",    "Наличные"),
        ("mkassa",  "Банк1"),
        ("zadatok", "Задаток"),
        ("optima",  "Банк2"),
    ]

    error = None
    if request.method == "POST":
        action  = request.POST.get("action")
        account = request.POST.get("account", "cash")
        comment = request.POST.get("comment", "")
        happened_at_str = request.POST.get("happened_at", "")
        try:
            amount = Decimal(request.POST.get("amount", "0"))
        except Exception:
            amount = Decimal("0")

        happened_at = None
        if happened_at_str:
            try:
                from datetime import datetime
                happened_at = timezone.make_aware(datetime.strptime(happened_at_str, "%Y-%m-%dT%H:%M"))
            except Exception:
                pass

        try:
            if action == "income":
                global_cash_income(account=account, amount=amount, comment=comment, created_by=request.user, happened_at=happened_at)
                messages.success(request, f"Пополнение {amount} добавлено в общую кассу.")
            elif action == "expense":
                global_cash_expense(account=account, amount=amount, comment=comment, created_by=request.user, happened_at=happened_at)
                messages.success(request, f"Расход {amount} распределён по точкам.")
            return redirect("dds:global_cash")
        except Exception as e:
            error = str(e)

    return render(request, "dds/global_cash.html", {
        "gcr": gcr,
        "ops": ops,
        "account_choices": ACCOUNT_CHOICES,
        "error": error,
    })


@login_required
def contact_add(request, hotel_id):
    from .models import PointContact
    from .forms import PointContactForm
    hotels_qs = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels_qs, pk=hotel_id)
    if request.method == "POST":
        form = PointContactForm(request.POST)
        if form.is_valid():
            contact = form.save(commit=False)
            contact.point = hotel
            contact.save()
    return redirect("dds:hotel_detail", pk=hotel_id)


@login_required
def contact_delete(request, hotel_id, contact_id):
    from .models import PointContact
    hotels_qs = user_hotels_qs(request.user)
    hotel = get_object_or_404(hotels_qs, pk=hotel_id)
    contact = get_object_or_404(PointContact, pk=contact_id, point=hotel)
    if request.method == "POST":
        contact.delete()
    return redirect("dds:hotel_detail", pk=hotel_id)
